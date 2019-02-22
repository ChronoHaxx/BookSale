from openpyxl import load_workbook
from prettytable import PrettyTable
import msvcrt as m

# REMAINING BUG
#When you go to action 1 or 2 then input 'Q', then action 3, it will not working properly..
#a lot of bug..


#wait for any key pressed
def wait():
    print("Press any key to continue...", end = "\r")
    m.getch()


#loading excel wb and declaring the sheet
excfile = 'DATA_BUKU.xlsx'
wb = load_workbook(filename = excfile)
sheet1 = wb['Sheet1']


#prompting for action
def promptAction():
    global action
    print("KEY                                                          \n1.INSERT TICK\n2.DELETE TICK\n3.VIEW ROW\n4.ANALYSIS\n5.QUICK TICK\n6.QUIT")
    action = input("Enter a key :")
    
    if action == '1' :
        print('You chose to insert a tick to a book row.')
        act12()
    elif action == '2' :
        print('You chose to delete a tick from a book row.')
        act12()
    elif action == '3' :
        print('You chose to view the row')
        act3()
    elif action == '4' :
        print("I'll give you an overview analysis\n NOT YET DONE SORRY ANA AND BWP :C")
    elif action == '5' :
        print("You chose quick tick, be careful... C: \nEnter 'Q' to quit.")
        act5()
    elif action == '6' :
        wb.save(filename = excfile)
        exit()
    else :
        print('Your input is out of context, please try again.')
        wait()
        promptAction()

#action 1 and 2
def act12():
    promptNo_siri()
    findCell(sheet1, no_siri)
    cellCheck(sheet1, i)
    tickCoord(sheet1, cell)
    if action == '1' :
        insertTick(sheet1, inputnew)
        newTable = PrettyTable()
        newTable.field_names = (list_with_values_header)
        list_with_values[5] = "/"
        list_with_values[4] = inputprice
        list_with_values[6] = inputbuyer
        newTable.add_row(list_with_values)
        print(newTable)
        wb.save(filename = excfile)
        wait()
        promptAction()
    elif action == '2' :
        deleteTick(sheet1, inputnew)
        newTable = PrettyTable()
        newTable.field_names = (list_with_values_header)
        list_with_values[5] = ""
        newTable.add_row(list_with_values)
        print(newTable)
        wb.save(filename = excfile)
        wait()
        promptAction()
    else :
        wb.save(filename = excfile)
        exit()
        
def act3():
    promptNo_siri()
    global i
    i = 1
    for row in sheet1.iter_rows():
        if sheet1.cell(row = i,column = 1).value == int(no_siri) :
            #   print(sheet1.cell(row = i,column = 1))
            #printing the excel first row (header)
            list_with_values_header=[]
            for cell in sheet1[1]:
                list_with_values_header.append(cell.value)
            #print(list_with_values_header)
            #printing the row
            list_with_values=[]
            for cell in sheet1[i]:
                list_with_values.append(cell.value)
            #print(list_with_values)
            #   print(sheet1.cell(row = i,column = 1).coordinate)

            #prettytable build table
            newTable = PrettyTable()
            newTable.field_names = (list_with_values_header)
            newTable.add_row(list_with_values)
            print(newTable)
            wait()

            break
        i += 1
        if i == 240 :
            print("There's no book with serial number of " + str(no_siri))    
            wait()

    

    
def act5():
    global act5_bool
    act5_bool = True
    while act5_bool :
            global i
            promptNo_siri()
            if act5_bool :
                findCell(sheet1, no_siri)
                cellCheck(sheet1, i)
                print(act5_bool)
                if i == 420 :
                    print("No book found, please check for typos or use another number.")
                    promptNo_siri()
                elif i > 420 : 
                    print("No book found, please check for typos or use another number.")
                    promptNo_siri()
                else:
                    tickCoord(sheet1, cell)
                    insertTick(sheet1, inputnew)
                    wb.save(filename = excfile)
                    newTable = PrettyTable()
                    newTable.field_names = (list_with_values_header)
                    list_with_values[5] = "/"
                    list_with_values[4] = inputprice
                    list_with_values[6] = inputbuyer
                    newTable.add_row(list_with_values)
                    print(newTable)
                    

#prompting for cell value
def promptNo_siri():
    global no_siri
    global act5_bool
    print("                                                                                            ")
    no_siri = input("No siri :")
    try:
        no_siri = int(no_siri)
        if no_siri < 0 :
            print("Only positive number please.")
            wait()
            promptNo_siri()
        else :
            True
    except ValueError:
        if no_siri == 'Q':
            act5_bool = False
            promptAction()
        else :
            print("That's not an int!")
            wait()
            promptNo_siri() 

#finding cell and its row
def findCell(sheet1, no_siri):
    global i, list_with_values_header, list_with_values
    i = 1
    for row in sheet1.iter_rows():
        if sheet1.cell(row = i,column = 1).value == int(no_siri) :
            #   print(sheet1.cell(row = i,column = 1))
            #printing the excel first row (header)
            list_with_values_header=[]
            for cell in sheet1[1]:
                list_with_values_header.append(cell.value)
                (list_with_values_header)
            #printing the row
            list_with_values=[]
            for cell in sheet1[i]:
                list_with_values.append(cell.value)
            #   print(list_with_values)
            #   print(sheet1.cell(row = i,column = 1).coordinate)
            
            break
        i += 1
    if i == 420 :
        print("No book found, please check for typos or use another number.")
        promptNo_siri()

#checking for correct value.
def cellCheck(sheet1, i):
    global cell
    cell = sheet1.cell(row = i,column = 1).coordinate
    #print("value for cell " + cell + ": " + str(sheet1[cell].value)) 

#list used to turn string to char array
#print(list(cell))

#changing to tick coord
def tickCoord(sheet1, cell):
    # row and col char are separated
    char1 = list(cell)[0]
    
    #this is a very unoptimized lazy fix (update: it broke tick again :C)
    #try :
    #    bool(list(cell)[3])
    #    char4 = list(cell)[3]
    #except:
    #    try:
    #        bool(list(cell)[2])
    #        char3 = list(cell)[2]
    #    except:
    #        bool(list(cell)[1])
    #        char2 = list(cell)[1]
    #    
    #try:
    #        bool(list(cell)[2])
    #        char3 = list(cell)[2]
    #except:
    #        bool(list(cell)[1])
    #        char2 = list(cell)[1]
    #
    #bool(list(cell)[1])
    #char2 = list(cell)[1]

    # this is first workaround but it return error for 1 digit no_siri
    #if bool(list(cell)[1]) == True:
    #    char2 = list(cell)[1]
    #elif bool(list(cell)[2]) == True:
    #    char3 = list(cell)[2]
    #elif bool(list(cell)[3]) == True:
    #    char4 = list(cell)[3]
    #else:

    #checking and incrementing it to be at 'Tick" column
    #   print(char1)
    charnew = chr(ord(char1) + 5)
    charnewprice = chr(ord(char1) + 4)
    charnewbuyer = chr(ord(char1) + 6)
    #   print(charnew)
    #joining existing constant row char with new col value
    global inputnew
    global inputnewprice
    global inputnewbuyer
    #another lazy fix
    #try :
    #    bool(list(cell)[3])
    #    inputnew = ''.join(charnew + char2 + char3 + char4)
    #except:
    #    try:
    #        bool(list(cell)[2])
    #        inputnew = ''.join(charnew + char2 + char3)
    #    except:
    #        bool(list(cell)[1])
    #        inputnew = ''.join(charnew + char2)
    #    
    #try:
    #        bool(list(cell)[2])
    #        inputnew = ''.join(charnew + char2 + char3)
    #except:
    #        bool(list(cell)[1])
    #        inputnew = ''.join(charnew + char2)
    #
    #bool(list(cell)[1])
    #inputnew = ''.join(charnew + char2)

    #
    inputnew = charnew
    inputnewprice = charnewprice
    inputnewbuyer = charnewbuyer
    for x in cell :
        if x == 'A':
            True
            #print("A is dismissed here")
        else:
            #print(x)
            new = "".join(x)
            inputnew += new

    for x in cell :
        if x == 'A':
            True
            #print("A is dismissed here")
        else:
            #print(x)
            new = "".join(x)
            inputnewprice += new

    for x in cell :
        if x == 'A':
            True
            #print("A is dismissed here")
        else:
            #print(x)
            new = "".join(x)
            inputnewbuyer += new

    #print(inputnew)
        

    #checking existing value at Tick column
    #print("Coord " + inputnew +" before : "+ str(sheet1[inputnew].value) )

def insertTick(sheet1, inputnew):
    #inserting a 'tick' ("/")
    global inputprice
    global inputbuyer
    inputbuyer = input("Who bought this book?\ne.g: 'HAZIMI' or 'MUHAMMAD' without '' marks\n")
    inputprice = input("How much did he pay for it? \ne.g: '3.00' or '0.00' without '' marks\n")
    sheet1[inputnew] = "/"
    sheet1[inputnewprice] = inputprice
    sheet1[inputnewbuyer] = inputbuyer
    #checking its value now
    print("Coord " + inputnew +" after : "+ str(sheet1[inputnew].value)  )
    

def deleteTick(sheet1, inputnew):
    #deleting a 'tick' ("/")
    sheet1[inputnew] = ""
    #checking its value now
    print("Coord " + inputnew +" after : "+ str(sheet1[inputnew].value)  )

while True :
    promptAction()


#saving the file
#make sure to close the excel or else perm is denied
wb.save(filename = excfile)
